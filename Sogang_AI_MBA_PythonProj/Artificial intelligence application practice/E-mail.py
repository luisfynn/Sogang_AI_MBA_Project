#이메일 전송 함수 코딩 (email_main.py)
import smtplib

smtp_info = dict({"smtp_server":"smtp.naver.com",
                  "smtp_user_id":"luisfynn@naver.com",
                  "smtp_user_pw":"@kj0224kj@",
                  "smtp_port":587})
######################################################################
def send_email(smtp_info, msg):
    with smtplib.SMTP(smtp_info["smtp_server"], smtp_info["smtp_port"]) as server:
        server.starttls() #TLS 보안 연결
        server.login(smtp_info["smtp_user_id"], smtp_info["smtp_user_pw"]) # 로그인
        # 로그인된 서버에 이메일 전송
        response = server.sendmail(msg['from'], msg['to'], msg.as_string())

        if not response:
            print('이메일을 성공적으로 보냈습니다.')
        else:
            print(response)

#이메일 내용 작성 및 파일 첨부 함수 코딩
import os
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase

#아래 msg_dict는 연습용
msg_dict = {
    'text': {'maintype': 'text', 'subtype': 'plain', 'filename': 'C://email/test1.txt'},
    'image': {'maintype': 'image', 'subtype': 'jpg', 'filename': 'C://email/test1.jpg'},
    'audio': {'maintype': 'audio', 'subtype': 'mp3', 'filename': 'C://email/test1.mp3'},
    'video': {'maintype': 'video', 'subtype': 'mp4', 'filename': 'C://email/test1.mp4'},
    'application': {'maintype': 'application', 'subtype': 'octect-stream',
                    'filename': 'C://email/test1.pdf'}

}

def make_multimsg(msg_dict):
    multi = MIMEMultipart(_subtype='mixed')
    for key, value in msg_dict.items():
        if key == 'text':
            with open(value['filename'], encoding='utf-8') as fp:
                msg = MIMEText(fp.read(), _subtype=value['subtype'])
        elif key == 'image':
            with open(value['filename'], 'rb') as fp:
                msg = MIMEImage(fp.read(), _subtype=value['subtype'])
        elif key == 'audio':
            with open(value['filename'], 'rb') as fp:
                msg = MIMEAudio(fp.read(), _subtype=value['subtype'])
        else:
            with open(value['filename'], 'rb') as fp:
                msg = MIMEBase(value['maintype'], _subtype=value['subtype'])
                msg.set_payload(fp.read())
                encoders.encode_base64(msg)
        msg.add_header('Content-Disposition', 'attachment', filename=os.path.basename(value['filename']))
        multi.attach(msg)
    return multi

#######################################################################
#실습1: 텍스트로만 구성된 이메일 전송
from email.mime.text import MIMEText # 이메일 내용 작성
title = '이메일 전송 자동화'
content = '이메일 자동화 전송입니다.'
sender = smtp_info['smtp_user_id']
receiver = 'luisfynn1@gmail.com'

msg = MIMEText(_text = content, _charset = 'utf-8')
msg['Subject'] = title
msg['From'] = sender
msg['To'] = receiver
# 이메일 전송
send_email(smtp_info, msg)
# 수신인 이메일 확인해 볼 것

#######################################################################
#실습2: 파일 첨부된 이메일 전송
title = '첨부화일 포함 이메일 전송 자동화'
content = '파일이 첨부된 이메일 자동화 전송입니다.'
sender = smtp_info['smtp_user_id']
receiver ='luisfynn1@gmail.com'

msg = MIMEText (_text = content, _charset = 'utf-8')

# 첨부파일 지정 (아래 경로에 해당 파일이 있어야 함)
msg_dict = {
    'text': {'maintype':'text','subtype':'plain','filename':'test1.txt'},
    'image' : {'maintype':'image','subtype':'jpg','filename':'test2.jpg'},
    'application': {'maintype':'application','subtype':'octect-stream','filename':'test3.pdf'}
}

# 첨부파일 추가
multi = make_multimsg(msg_dict) # msg_dict dict안의 파일을 첨부함
multi['Subject'] = title
multi['From'] = sender
multi['To'] = receiver
multi.attach(msg)
# 이메일 전송
send_email(smtp_info, multi)
# 수신인 이메일 확인해 볼 것

#######################################################################
#실습3: 특정 사이트에서 이메일 형식 자료 수집하여 엑셀에 기록하기(email_send.py)
import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook
url = 'https://news.v.daum.net/v/20211129144552297'
headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

response = requests.get(url, headers=headers)
results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text) #정규 표현식. response.text 내에서 email 주소를 찾는 패턴
response.text
results #output : ['wine_sky@news1.kr', 'Y6z3VwK78FOFN59061b64xYu@timetalk', 'wine_sky@news1.kr']
results = list(set(results))
print(results)

try:
    wb = load_workbook("email_list.xlsx", data_only = True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])
wb.save(r"email_list.xlsx")

#######################################################################
#실습4: 엑셀 파일 안의 이메일 주소 읽어서 대량 이메일 전송
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

load_wb = load_workbook(r"email_list.xlsx", data_only = True)
load_ws = load_wb.active

for i in range(1,load_ws.max_row + 1):
    recv_email_value = load_ws.cell(i, 1).value
    print("성공:",recv_email_value)
    try:
        send_email = "luisfynn@naver.com"  # 네이버 이메일 주소
        send_pwd = "@kj0224kj@"    # 네이버 비밀번호
        recv_email = recv_email_value
        smtp_name = "smtp.naver.com"
        smtp_port = 587
        msg = MIMEMultipart()
        msg['Subject'] ="엑셀에서 메일주소를 읽어 자동으로 보내는 메일입니다."
        msg['From'] = send_email
        msg['To'] = recv_email
        text = """
            메일내용 입니다. 
            감사합니다.
            """
        msg.attach( MIMEText(text) )

        s=smtplib.SMTP(smtp_name,smtp_port)
        s.starttls()
        s.login(send_email,send_pwd)
        s.sendmail( send_email, recv_email, msg.as_string() )
        s.quit()
    except:
        print("에러:",recv_email_value)
